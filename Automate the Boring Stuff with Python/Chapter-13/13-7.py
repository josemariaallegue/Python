import openpyxl
from openpyxl.styles import Font
from pathlib import Path

print("Excel - merge y unmerge")
path = Path(
    r"C:\Users\jmallegue\Documents\GitHub\Python\Automate the Boring Stuff with Python\Archivos auxiliares\Chapter-13\Ejemplo1.xlsx")

wbExample = openpyxl.load_workbook(path)
wsSheet1 = wbExample["Prueba2"]

# merge
wsSheet1.merge_cells('A1:D3')  # Merge all these cells.
wsSheet1['A1'] = 'Twelve cells merged together.'
wsSheet1.merge_cells('C5:D5')  # Merge these two cells.
wsSheet1['C5'] = 'Two merged cells.'

# unmerge
# wsSheet1.unmerge_cells('A1:D3')  # Split these cells up.
# wsSheet1.unmerge_cells('C5:D5')

wbExample.save(path)
