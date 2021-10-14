import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from pathlib import Path

print("Excel - propiedades utiles y datos de un rango")
path = Path(
    r"C:\Users\jmallegue\Documents\GitHub\Python\Automate the Boring Stuff with Python\Archivos auxiliares\Chapter-13\example.xlsx")

wbExample = openpyxl.load_workbook(path)
wsSheet1 = wbExample["Sheet1"]

print(wsSheet1.max_column)  # obetner cantidad de columnas
print(wsSheet1.max_row)  # obtener cantidad de filas

# convertir un numero a letra - para la ubicacion de columnas
print(get_column_letter(1))
# convertir una letra a numero - para la ubicacion de columnas
print(column_index_from_string("a"))

# siguiente ejemplo permite recorrer un rango de celdas de top-left a bottom-right
for rowOfCellObjects in wsSheet1['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
        print('--- END OF ROW ---')

# para acceder a los datos de una fila o columna en particula
for cellObj in list(wsSheet1.columns)[1]:
    print(cellObj.value)
