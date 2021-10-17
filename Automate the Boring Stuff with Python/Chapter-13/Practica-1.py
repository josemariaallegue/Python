from types import MappingProxyType
import openpyxl
from pathlib import Path
import pyinputplus as pyip

path = Path(r"D:\Documentos\GitHub\Python\Automate the Boring Stuff with Python\Archivos auxiliares\Chapter-13", "Practica-1.xlsx")

# revision de existencia
if(path.exists()):
    wbAux = openpyxl.load_workbook(path)
    wsAux = wbAux["Sheet1"]

else:
    wbAux = openpyxl.Workbook()
    wsAux = wbAux("Sheet")
    wbAux.save(path)

# input y calculo
maximo = pyip.inputNum("Por favor ingrese un numero: ", default=1)

for i in range(1, maximo+2):
    wsAux.cell(1, i).value = i-1
    wsAux.cell(i, 1).value = i-1

for i in range(2, maximo+2):
    for j in range(2, maximo+2):
        wsAux.cell(i, j).value = (i-1)*(j-1)
        print(str((i-1)*(j-1)))

wbAux.save(path)
