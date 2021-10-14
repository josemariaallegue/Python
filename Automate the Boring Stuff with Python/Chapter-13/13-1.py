import openpyxl
from pathlib import Path

print("Excel - conceptos basicos")
path = Path(
    r"C:\Users\jmallegue\Documents\GitHub\Python\Automate the Boring Stuff with Python\Archivos auxiliares\Chapter-13\example.xlsx")

wbExample = openpyxl.load_workbook(path)

for sheet in wbExample.sheetnames:
    print(sheet)

wsHoja = wbExample["Sheet1"]  # obtencion de hoja
celda = wsHoja["A1"]  # obtencion de celda
# atributos
print(f"Valor: {celda.value}, columna: {celda.column}, fila: {celda.row}, coordenadas: {celda.coordinate}")

# forma de recorrer celdas
for i in range(1, 8, 2):
    # forma de acceder a una celda a travez de la columna y fila
    print(i, wsHoja.cell(row=i, column=2).value)
